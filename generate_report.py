from __future__ import annotations

import json
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
INPUT_DIR = ROOT / "01_ОСВ_входящие"
OUTPUT_DIR = ROOT / "02_Готовые_отчеты"
TEMPLATE_DIR = ROOT / "03_Шаблоны"
CONFIG_PATH = ROOT / "04_Настройки" / "report_lines.json"
TEMPLATE_PATH = TEMPLATE_DIR / "Шаблон отчета.xlsx"


@dataclass
class OsvRow:
    account: str
    name: str
    start_debit: float
    start_credit: float
    turnover_debit: float
    turnover_credit: float
    end_debit: float
    end_credit: float


def parse_money(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text or text in {"-", "—"}:
        return 0.0

    is_negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    text = text.replace("\u00a0", " ").replace(" ", "")
    text = text.replace(",", ".")
    text = re.sub(r"[^0-9.\-]", "", text)
    if text in {"", "-", ".", "-."}:
        return 0.0

    number = float(text)
    return -number if is_negative else number


def is_money_cell(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True

    text = str(value).strip()
    if not text or text in {"-", "—"}:
        return True

    return bool(re.match(r"^[\s\u00a0()\-—0-9.,]+$", text))


def is_account(value: Any) -> bool:
    text = str(value or "").strip()
    return bool(re.match(r"^\d{2}(?:[.\-/]\d+)*$", text))


def normalize_account(value: Any) -> str:
    return str(value or "").strip().replace(",", ".")


def split_account_and_name(value: Any) -> tuple[str, str]:
    text = str(value or "").strip()
    match = re.match(r"^(\d{2}(?:[.\-/]\d+)*)(?:\s*,\s*(.*))?$", text)
    if not match:
        return normalize_account(text), ""
    return normalize_account(match.group(1)), (match.group(2) or "").strip()


def find_header_row(ws) -> tuple[int, int] | None:
    for row_idx in range(1, min(ws.max_row, 100) + 1):
        for col_idx in range(1, min(ws.max_column, 30) + 1):
            value = ws.cell(row_idx, col_idx).value
            text = str(value or "").strip().lower()
            if text in {"счет", "счёт"} or text.startswith(("счет,", "счёт,")):
                return row_idx, col_idx
    return None


def collect_metadata(ws) -> dict[str, str]:
    metadata = {"source_sheet": ws.title, "period": "", "organization": ""}

    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 25), values_only=True):
        line = " ".join(str(cell).strip() for cell in row if cell not in (None, ""))
        lowered = line.lower()
        if not metadata["period"] and "оборотно-сальдовая ведомость" in lowered:
            metadata["period"] = line
        if not metadata["period"] and "период" in lowered and "сальдо на начало периода" not in lowered:
            metadata["period"] = line
        if not metadata["organization"] and ("организация" in lowered or "по организации" in lowered):
            metadata["organization"] = line

    return metadata


def parse_osv(path: Path) -> tuple[list[OsvRow], dict[str, str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    header = find_header_row(ws)

    if header:
        start_row, account_col = header[0] + 1, header[1]
    else:
        start_row, account_col = 1, 1

    rows: list[OsvRow] = []
    for row_idx in range(start_row, ws.max_row + 1):
        account, embedded_name = split_account_and_name(ws.cell(row_idx, account_col).value)
        if not is_account(account):
            continue

        trailing_values = [ws.cell(row_idx, col).value for col in range(account_col + 1, ws.max_column + 1)]
        money_values = trailing_values[:6]

        if len(money_values) < 6:
            continue

        name_parts: list[str] = []
        for value in trailing_values[6:]:
            if value in (None, ""):
                continue
            if isinstance(value, str) and not re.search(r"\d", value):
                name_parts.append(value.strip())
            if len(name_parts) >= 2:
                break

        start_debit, start_credit, turnover_debit, turnover_credit, end_debit, end_credit = [
            parse_money(value) for value in money_values
        ]
        rows.append(
            OsvRow(
                account=account,
                name=embedded_name or " ".join(name_parts),
                start_debit=start_debit,
                start_credit=start_credit,
                turnover_debit=turnover_debit,
                turnover_credit=turnover_credit,
                end_debit=end_debit,
                end_credit=end_credit,
            )
        )

    if not rows:
        raise ValueError(
            "Не удалось найти строки ОСВ. Проверьте, что файл .xlsx содержит колонку 'Счет' "
            "и 6 денежных колонок: начальное сальдо Дт/Кт, обороты Дт/Кт, конечное сальдо Дт/Кт."
        )

    return rows, collect_metadata(ws)


def load_config() -> dict[str, Any]:
    with CONFIG_PATH.open("r", encoding="utf-8") as file:
        return json.load(file)


def account_matches(account: str, prefixes: list[str]) -> bool:
    normalized = account.replace(",", ".")
    return any(normalized == prefix or normalized.startswith(f"{prefix}.") for prefix in prefixes)


def value_for_mode(row: OsvRow, mode: str) -> float:
    if mode == "end_debit":
        return row.end_debit
    if mode == "end_credit":
        return row.end_credit
    if mode == "end_debit_net":
        return row.end_debit - row.end_credit
    if mode == "end_credit_net":
        return row.end_credit - row.end_debit
    if mode == "turnover_debit":
        return row.turnover_debit
    if mode == "turnover_credit":
        return row.turnover_credit
    if mode == "turnover_credit_minus_debit":
        return row.turnover_credit - row.turnover_debit
    if mode == "turnover_debit_minus_credit":
        return row.turnover_debit - row.turnover_credit
    raise ValueError(f"Неизвестный режим расчета: {mode}")


def calculate(rows: list[OsvRow], prefixes: list[str], mode: str) -> float:
    total = 0.0
    by_account = {row.account.replace(",", "."): row for row in rows}
    side_specific_modes = {"end_debit", "end_credit"}

    for prefix in prefixes:
        normalized_prefix = prefix.replace(",", ".")

        if mode in side_specific_modes:
            descendants = [
                row for row in rows if row.account.replace(",", ".").startswith(f"{normalized_prefix}.")
            ]
            if descendants:
                descendant_accounts = [row.account.replace(",", ".") for row in descendants]
                leaf_rows = [
                    row
                    for row in descendants
                    if not any(
                        account.startswith(f"{row.account.replace(',', '.')}.")
                        for account in descendant_accounts
                    )
                ]
                total += sum(value_for_mode(row, mode) for row in leaf_rows)
                continue

        exact_row = by_account.get(normalized_prefix)
        if exact_row:
            total += value_for_mode(exact_row, mode)
            continue

        for row in rows:
            if row.account.replace(",", ".").startswith(f"{normalized_prefix}."):
                total += value_for_mode(row, mode)

    return total


def style_sheet(ws) -> None:
    thin_gray = Side(style="thin", color="D9DEE8")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="EAF2F8")
    total_fill = PatternFill("solid", fgColor="D9EAD3")

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx in range(2, ws.max_row + 1):
        label = str(ws.cell(row_idx, 1).value or "")
        if label.startswith("Раздел:"):
            for cell in ws[row_idx]:
                cell.font = Font(bold=True)
                cell.fill = section_fill
        if label.startswith("Итого") or label.startswith("Контроль"):
            for cell in ws[row_idx]:
                cell.font = Font(bold=True)
                cell.fill = total_fill

    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for cell in ws[get_column_letter(col_idx)]:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 45)


def create_template() -> None:
    TEMPLATE_DIR.mkdir(exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет"
    ws.append(["Показатель", "Сумма", "Комментарий"])
    ws.append(["Раздел: Активы", "", ""])
    ws.append(["Денежные средства", "", "50, 51, 52, 55"])
    ws.append(["Авансы поставщикам", "", "60.02 по дебету"])
    ws.append(["Дебиторская задолженность покупателей", "", "62.01 по дебету"])
    ws.append(["Подотчетные лица", "", "71 по дебету"])
    ws.append(["Переплаты по налогам", "", "68 по дебету"])
    ws.append(["Переплаты по страховым взносам", "", "69 по дебету"])
    ws.append(["Запасы, товары и незавершенное производство", "", "10, 20, 41, 43, 44"])
    ws.append(["Внеоборотные активы", "", "01, 03, 04, 08 минус амортизация 02, 05"])
    ws.append(["Итого активы", "", ""])
    ws.append(["Раздел: Обязательства", "", ""])
    ws.append(["Задолженность перед поставщиками", "", "60.01 по кредиту"])
    ws.append(["Авансы покупателей", "", "62.02 по кредиту"])
    ws.append(["Налоги к уплате", "", "68 по кредиту"])
    ws.append(["Страховые взносы к уплате", "", "69 по кредиту"])
    ws.append(["Кредиты и займы", "", "66, 67"])
    ws.append(["Итого обязательства", "", ""])
    ws.append(["Раздел: Собственный капитал", "", ""])
    ws.append(["Уставный капитал", "", "80 по кредиту"])
    ws.append(["Накопленная прибыль / убыток", "", "Активы минус обязательства минус уставный капитал"])
    ws.append(["Итого собственный капитал", "", ""])
    ws.append(["Итого пассивы и капитал", "", ""])
    ws.append(["Контроль: активы - пассивы и капитал", "", "Должно быть 0"])
    style_sheet(ws)
    wb.save(TEMPLATE_PATH)


def write_report(source_path: Path, rows: list[OsvRow], metadata: dict[str, str], config: dict[str, Any]) -> Path:
    OUTPUT_DIR.mkdir(exist_ok=True)
    create_template()

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет"
    ws.append([config.get("report_title", "Управленческий баланс по ОСВ"), "", ""])
    ws.append(["Источник", source_path.name, ""])
    ws.append(["Дата формирования", datetime.now().strftime("%d.%m.%Y %H:%M"), ""])
    ws.append(["Период", metadata.get("period", ""), ""])
    ws.append(["Организация", metadata.get("organization", ""), ""])
    ws.append(["", "", ""])
    ws.append(["Показатель", "Сумма", "Комментарий"])

    section_totals: dict[str, float] = {}
    calculated_values: dict[str, float] = {}
    current_section = ""
    for section in config["sections"]:
        current_section = section["name"]
        section_totals[current_section] = 0.0
        ws.append([f"Раздел: {current_section}", "", section.get("comment", "")])

        for line in section["lines"]:
            if line["mode"] == "balance_to_assets":
                value = section_totals.get("Активы", 0.0) - section_totals.get("Обязательства", 0.0)
            elif line["mode"] == "equity_remainder":
                equity_before_result = section_totals[current_section]
                value = (
                    section_totals.get("Активы", 0.0)
                    - section_totals.get("Обязательства", 0.0)
                    - equity_before_result
                )
            else:
                value = calculate(rows, line["accounts"], line["mode"])
            section_totals[current_section] += value
            calculated_values[line["title"]] = value
            ws.append([line["title"], value, line.get("comment", "")])

        if section.get("show_total", True):
            ws.append([f"Итого {current_section.lower()}", section_totals[current_section], ""])

    if "Активы" in section_totals and "Обязательства" in section_totals and "Собственный капитал" in section_totals:
        liabilities_and_equity = section_totals["Обязательства"] + section_totals["Собственный капитал"]
        control = section_totals["Активы"] - liabilities_and_equity
        ws.append(["Итого пассивы и капитал", liabilities_and_equity, "Обязательства + собственный капитал"])
        ws.append(["Контроль: активы - пассивы и капитал", control, "Должно быть 0"])

    capital_calc = wb.create_sheet("Расчет капитала")
    assets = section_totals.get("Активы", 0.0)
    liabilities = section_totals.get("Обязательства", 0.0)
    charter_capital = calculated_values.get("Уставный капитал", 0.0)
    retained_result = calculated_values.get("Накопленная прибыль / убыток", 0.0)
    equity = section_totals.get("Собственный капитал", 0.0)
    capital_calc.append(["Расчет", "Сумма", "Пояснение"])
    capital_calc.append(["Активы", assets, "Итого активы управленческого баланса"])
    capital_calc.append(["Минус обязательства", -liabilities, "Итого обязательства"])
    capital_calc.append(["Собственный капитал", equity, "Активы - обязательства"])
    capital_calc.append(["Минус уставный капитал", -charter_capital, "Счет 80"])
    capital_calc.append(["Накопленная прибыль / убыток", retained_result, "Собственный капитал - уставный капитал"])
    capital_calc.append(["Контроль", assets - liabilities - charter_capital - retained_result, "Должно быть 0"])
    for row in capital_calc.iter_rows(min_row=2, max_row=capital_calc.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '#,##0.00;[Red]-#,##0.00;"-"'
    style_sheet(capital_calc)

    for row in ws.iter_rows(min_row=1, max_row=6):
        for cell in row:
            cell.border = Border()
            cell.font = Font(bold=cell.row == 1)

    for row in ws.iter_rows(min_row=7):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00;[Red]-#,##0.00;"-"'

    style_sheet(ws)
    ws.freeze_panes = "A8"

    detail = wb.create_sheet("Детализация ОСВ")
    detail.append(
        [
            "Счет",
            "Наименование",
            "Сальдо нач. Дт",
            "Сальдо нач. Кт",
            "Оборот Дт",
            "Оборот Кт",
            "Сальдо кон. Дт",
            "Сальдо кон. Кт",
        ]
    )
    for row in rows:
        detail.append(
            [
                row.account,
                row.name,
                row.start_debit,
                row.start_credit,
                row.turnover_debit,
                row.turnover_credit,
                row.end_debit,
                row.end_credit,
            ]
        )
    style_sheet(detail)

    mapping = wb.create_sheet("Настройка строк")
    mapping.append(["Раздел", "Строка", "Счета", "Режим", "Комментарий"])
    for section in config["sections"]:
        for line in section["lines"]:
            mapping.append(
                [
                    section["name"],
                    line["title"],
                    ", ".join(line["accounts"]),
                    line["mode"],
                    line.get("comment", ""),
                ]
            )
    style_sheet(mapping)

    safe_name = source_path.stem.replace(" ", "_")
    output_path = OUTPUT_DIR / f"Отчет_по_ОСВ_{safe_name}_{datetime.now():%Y%m%d_%H%M}.xlsx"
    wb.save(output_path)
    return output_path


def find_latest_osv() -> Path:
    candidates = sorted(INPUT_DIR.glob("*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)
    candidates = [path for path in candidates if not path.name.startswith("~$")]
    if not candidates:
        raise FileNotFoundError(f"Положите файл ОСВ .xlsx в папку: {INPUT_DIR}")
    return candidates[0]


def main() -> None:
    config = load_config()
    source_path = find_latest_osv()
    rows, metadata = parse_osv(source_path)
    output_path = write_report(source_path, rows, metadata, config)
    print(f"Готово: {output_path}")


if __name__ == "__main__":
    main()
